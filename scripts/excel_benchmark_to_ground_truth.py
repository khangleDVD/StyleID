"""
Chuyển file Excel benchmark (Benchmark Anh AI.xlsx) sang ground_truth_parsed.jsonl
để dùng với eval_fashion_accuracy.py.

Cột Excel kỳ vọng:
  Image_ID, Overall_Style, Item, Item_Style, Image

- Mỗi Image_ID có nhiều dòng (một dòng / món).
- Ảnh được nhúng trong cột Image (openpyxl), không phải đường dẫn text.
- Image_ID có thể merge cell: script tự forward-fill theo thứ tự dòng.

Cách dùng
---------
python scripts/excel_benchmark_to_ground_truth.py \
    --xlsx "Benchmark Anh AI.xlsx" \
    --out-dir artifacts/eval_excel

python eval_fashion_accuracy.py \
    --gt artifacts/eval_excel/ground_truth_parsed.jsonl \
    --media artifacts/eval_excel/media \
    --api http://127.0.0.1:5000 \
    --user-id 1 \
    --out artifacts/eval_excel/results
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit(
        "Thiếu openpyxl. Cài: pip install openpyxl"
    ) from exc


_ITEM_NUM_PREFIX = re.compile(r"^\d+\.\s*")


def _clean_item_name(name: str) -> str:
    s = (name or "").strip()
    s = _ITEM_NUM_PREFIX.sub("", s)
    return s.strip()


def _first_non_empty(values: List[Optional[str]]) -> str:
    for v in values:
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _mode_style(styles: List[str]) -> str:
    styles = [s.strip() for s in styles if s and str(s).strip()]
    if not styles:
        return ""
    return Counter(styles).most_common(1)[0][0]


@dataclass
class BenchmarkSample:
    image_id: str
    overall_style: str
    items: List[Tuple[str, str]] = field(default_factory=list)  # (item, item_style)
    image_path: Optional[Path] = None


def _read_rows(xlsx_path: Path, sheet: Optional[str]) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]

    header = [str(c or "").strip() for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    col_map = {name.lower(): idx for idx, name in enumerate(header)}

    def col(*names: str) -> int:
        for n in names:
            key = n.lower()
            if key in col_map:
                return col_map[key]
        raise KeyError(f"Không tìm thấy cột {names!r} trong header: {header}")

    i_id = col("image_id")
    i_overall = col("overall_style")
    i_item = col("item")
    i_style = col("item_style")

    rows: List[Dict[str, Any]] = []
    current_id = ""
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        raw_id = excel_row[i_id] if i_id < len(excel_row) else None
        if raw_id is not None and str(raw_id).strip():
            current_id = str(raw_id).strip()

        item = excel_row[i_item] if i_item < len(excel_row) else None
        if item is None or not str(item).strip():
            continue
        if not current_id:
            continue

        rows.append(
            {
                "image_id": current_id,
                "overall_style": excel_row[i_overall] if i_overall < len(excel_row) else None,
                "item": str(item).strip(),
                "item_style": excel_row[i_style] if i_style < len(excel_row) else None,
            }
        )

    wb.close()
    return rows


def _group_samples(rows: List[Dict[str, Any]]) -> List[BenchmarkSample]:
    grouped: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    order: List[str] = []
    for r in rows:
        iid = r["image_id"]
        if iid not in grouped:
            order.append(iid)
        grouped[iid].append(r)

    samples: List[BenchmarkSample] = []
    for iid in order:
        chunk = grouped[iid]
        overall = _first_non_empty([r.get("overall_style") for r in chunk])
        item_styles: List[str] = []
        items: List[Tuple[str, str]] = []
        for r in chunk:
            item = _clean_item_name(str(r.get("item") or ""))
            style = str(r.get("item_style") or "").strip()
            if not item:
                continue
            items.append((item, style))
            if style:
                item_styles.append(style)
        if not overall:
            overall = _mode_style(item_styles)
        samples.append(BenchmarkSample(image_id=iid, overall_style=overall, items=items))
    return samples


def _extract_embedded_images(xlsx_path: Path, media_dir: Path) -> List[Tuple[int, bytes]]:
    """Trả về list (excel_row_1based, image_bytes) sorted theo dòng."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[wb.sheetnames[0]]
    out: List[Tuple[int, bytes]] = []
    for img in ws._images:
        row_1based = img.anchor._from.row + 1
        out.append((row_1based, img._data()))
    wb.close()
    out.sort(key=lambda x: x[0])
    return out


def _guess_ext(data: bytes) -> str:
    if data[:8] == b"\x89PNG\r\n\x1a\n":
        return ".png"
    if data[:2] == b"\xff\xd8":
        return ".jpg"
    if data[:4] == b"RIFF" and data[8:12] == b"WEBP":
        return ".webp"
    return ".png"


def _attach_images(
    samples: List[BenchmarkSample],
    xlsx_path: Path,
    media_dir: Path,
) -> None:
    media_dir.mkdir(parents=True, exist_ok=True)
    for p in media_dir.iterdir():
        if p.is_file():
            p.unlink()

    images = _extract_embedded_images(xlsx_path, media_dir)
    if len(images) != len(samples):
        print(
            f"[warn] Số ảnh nhúng ({len(images)}) != số Image_ID ({len(samples)}). "
            "Ghép theo thứ tự xuất hiện; kiểm tra lại file Excel nếu lệch.",
            file=sys.stderr,
        )

    n = min(len(images), len(samples))
    for i in range(n):
        _row, data = images[i]
        ext = _guess_ext(data)
        out_path = media_dir / f"{samples[i].image_id}{ext}"
        out_path.write_bytes(data)
        samples[i].image_path = out_path.resolve()


def _build_gt_items_raw(items: List[Tuple[str, str]]) -> str:
    parts: List[str] = []
    for item, style in items:
        if style:
            parts.append(f"{item} - {style}")
        else:
            parts.append(item)
    return ", ".join(parts)


def _to_jsonl_record(idx: int, sample: BenchmarkSample) -> Dict[str, Any]:
    return {
        "sample_id": f"{idx:04d}_{sample.image_id}",
        "image_id": sample.image_id,
        "image_file": str(sample.image_path) if sample.image_path else "",
        "gt_items_raw": _build_gt_items_raw(sample.items),
        "gt_style_raw": "",
        "gt_overall_raw": sample.overall_style,
        "gt_item_count": len(sample.items),
    }


def convert_excel_benchmark(
    xlsx_path: Path,
    out_dir: Path,
    sheet: Optional[str] = None,
) -> Path:
    xlsx_path = xlsx_path.expanduser().resolve()
    out_dir = out_dir.resolve()
    media_dir = out_dir / "media"
    out_dir.mkdir(parents=True, exist_ok=True)

    rows = _read_rows(xlsx_path, sheet)
    if not rows:
        raise SystemExit(f"Không đọc được dòng dữ liệu nào từ {xlsx_path}")

    samples = _group_samples(rows)
    _attach_images(samples, xlsx_path, media_dir)

    gt_path = out_dir / "ground_truth_parsed.jsonl"
    with gt_path.open("w", encoding="utf-8") as f:
        for idx, sample in enumerate(samples, start=1):
            rec = _to_jsonl_record(idx, sample)
            f.write(json.dumps(rec, ensure_ascii=False) + "\n")

    summary = {
        "source_xlsx": str(xlsx_path),
        "num_samples": len(samples),
        "ground_truth_jsonl": str(gt_path),
        "media_dir": str(media_dir),
    }
    (out_dir / "manifest.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    return gt_path


def main() -> None:
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="backslashreplace")
    except Exception:
        pass

    ap = argparse.ArgumentParser(
        description="Chuyển Benchmark Anh AI.xlsx → ground_truth_parsed.jsonl + ảnh trong media/"
    )
    ap.add_argument(
        "--xlsx",
        default="Benchmark Anh AI.xlsx",
        help="Đường dẫn file Excel benchmark",
    )
    ap.add_argument(
        "--sheet",
        default=None,
        help="Tên sheet (mặc định: sheet đầu tiên)",
    )
    ap.add_argument(
        "--out-dir",
        default="artifacts/eval_excel",
        help="Thư mục output (media/ + ground_truth_parsed.jsonl)",
    )
    args = ap.parse_args()

    gt_path = convert_excel_benchmark(
        Path(args.xlsx),
        Path(args.out_dir),
        sheet=args.sheet,
    )
    print(f"Đã ghi ground truth: {gt_path}")
    print(f"Ảnh: {Path(args.out_dir).resolve() / 'media'}")
    print()
    print("Chạy eval:")
    print(
        "  python eval_fashion_accuracy.py "
        f'--gt "{gt_path}" '
        f'--media "{Path(args.out_dir).resolve() / "media"}" '
        "--api http://127.0.0.1:5000 --user-id 1 "
        f'--out "{Path(args.out_dir).resolve() / "results"}"'
    )


if __name__ == "__main__":
    main()
