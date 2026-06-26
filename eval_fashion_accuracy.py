"""
eval_fashion_accuracy.py
========================
Script kiểm chứng độ chính xác hệ thống nhận dạng thời trang
dựa trên ground_truth_parsed.jsonl đã có sẵn.

Cách dùng
---------
# Cần server Flask đang chạy (python app.py), MySQL, user trong DB và OPENROUTER_API_KEY (đánh giá thật).
# API /api/analyze bắt buộc user_id — dùng tài khoản admin (không trừ lượt) hoặc user đủ analysis_credits.
# Eval gửi skip_history=1 — không ghi bảng history (kết quả chỉ lưu trong artifacts/.../results).
# Tạo ground truth từ Excel (lần đầu hoặc sau khi cập nhật Benchmark Anh AI.xlsx):
python scripts/excel_benchmark_to_ground_truth.py \
    --xlsx "Benchmark Anh AI.xlsx" \
    --out-dir artifacts/eval_excel

python eval_fashion_accuracy.py \
    --gt artifacts/eval_excel/ground_truth_parsed.jsonl \
    --media artifacts/eval_excel/media \
    --api http://127.0.0.1:5000 \
    --user-id 1 \
    --out artifacts/eval_excel/results

# Tùy chọn thêm:
    --limit 10          # chỉ chạy 10 ảnh đầu (test nhanh)
    --sleep 2           # nghỉ 2 giây giữa các lần gọi API (tránh rate limit)
    --fuzzy-threshold 0.6   # ngưỡng similarity fuzzy match item (0..1)

Metric được tính
----------------
- Item detection F1 (per sample + aggregate):
    precision = |items_đúng| / |items_dự đoán|
    recall    = |items_đúng| / |items_ground_truth|
    F1        = harmonic mean(P, R)
  "Đúng" = fuzzy match (normalized edit distance) vượt ngưỡng

- Per-item style accuracy:
    Với mỗi item khớp được: style AI có khớp style GT không?
    (normalized, alias-aware)

- Tổng số lượng (toàn bộ benchmark):
    items_correct_total       — tổng item nhận diện đúng (TP, fuzzy match)
    items_style_correct_total — trong các TP, số item có style_id khớp GT
    gt_items_total / pred_items_total — tổng món GT và dự đoán (để đối chiếu)

- Overall style accuracy:
    GT: **gt_overall_raw** (nếu có) → chuẩn hoá; không thì **mode(style_id)** trên các món GT.
    Pred: **mode(style_id)** hoặc **overall_style** API (`--overall-match`).

- Outfit (aggregate_styles) — từ API `overall_style_top_k`:
    **overall_top1_accuracy**: GT có trùng hạng 1 trong bảng điểm outfit không.
    **overall_top3_accuracy**: GT có trùng một trong 3 hạng đầu không (`--overall-soft` áp dụng).
    **overall_macro_f1**: Macro F1 đa lớp, mỗi ảnh một dự đoán Top-1 (cùng cột như trên).

Output
------
- report.jsonl / report_<run_id>.jsonl   : chi tiết từng sample (bản latest + archive)
- report.xlsx / report_<run_id>.xlsx   : Excel (sheet Chi tiết + Tổng hợp)
- report.csv / report_<run_id>.csv     : một dòng/sample; UTF-8 BOM
- summary.json / summary_<run_id>.json : metric tổng hợp (`run_id` + `saved_at` trong JSON)
- errors.txt / errors_<run_id>.txt     : log lỗi API (có thể rỗng)
"""

import argparse
import csv
import json
import os
import re
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore[misc, assignment]
    Alignment = Font = get_column_letter = None  # type: ignore[misc, assignment]
import time
from datetime import datetime
from collections import Counter
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests

# ---------------------------------------------------------------------------
# §1 — Chuẩn hóa chuỗi
# ---------------------------------------------------------------------------

def _norm(s: str) -> str:
    """Chuẩn hóa lowercase, bỏ ký tự đặc biệt, gộp space."""
    s = (s or "").strip().lower()
    s = re.sub(r"[\u2013\u2014\u2012]", "-", s)
    s = re.sub(r"[^\w\s\-\/]+", "", s, flags=re.UNICODE)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def _fuzzy(a: str, b: str) -> float:
    """Similarity [0..1] giữa 2 chuỗi đã normalize."""
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    return SequenceMatcher(None, a, b).ratio()


# ---------------------------------------------------------------------------
# §2 — Style normalization: map tên GT → style_id trong dataset
# ---------------------------------------------------------------------------

# Map thủ công các style GT hay xuất hiện nhưng không có trong dataset,
# hoặc có alias không rõ. Dựa trên 49 style trong fashion_style_dataset.json.
_STYLE_ALIAS: Dict[str, str] = {
    # GT dùng tên đầy đủ, dataset dùng id lowercase
    "streetwear": "streetwear",
    "hip hop": "hip_hop",
    "hip-hop": "hip_hop",
    "hiphop": "hip_hop",
    "techwear": "techwear",
    "athleisure": "athleisure",
    "skater style": "skater_style",
    "skater": "skater_style",
    "hypebeast": "hypebeast",
    "classic": "classic",
    "minimalist": "minimalist",
    "minimalism": "minimalist",
    "minimal": "minimalist",
    "formal": "formal",
    "business": "business",
    "smart casual": "smart_casual",
    "preppy": "preppy",
    "old money": "old_money",
    "vintage": "vintage",
    "retro": "retro",
    "grunge": "grunge",
    "punk": "punk",
    "gothic": "gothic",
    "goth": "gothic",
    "bohemian": "bohemian",
    "boho": "bohemian",
    "artsy": "artsy",
    "feminine": "feminine",
    "sexy": "sexy",
    "elegant": "elegant",
    "elegance": "elegant",
    "classic feminine": "feminine",
    "minimal casual": "minimalist",
    "soft girl": "soft_girl",
    "coquette": "coquette",
    "casual": "casual",
    "normcore": "normcore",
    "basic": "basic",
    "comfy": "comfy",
    "sporty": "sporty",
    "activewear": "activewear",
    "blokecore": "blokecore",
    "korean style": "korean_style",
    "k-fashion": "korean_style",
    "k fashion": "korean_style",
    "japanese streetwear": "japanese_streetwear",
    "harajuku": "harajuku",
    "chinese street style": "chinese_street_style",
    "parisian style": "parisian_style",
    "parisian": "parisian_style",
    "italian style": "italian_style",
    "cultural dress": "cultural",
    "traditional": "cultural",
    "ethnic wear": "cultural",
    "uniform": "uniform",
    "dark academia": "dark_academia",
    "light academia": "light_academia",
    "y2k": "y2k",
    "cyberpunk": "cyberpunk",
    "egirl": "egirl_eboy",
    "eboy": "egirl_eboy",
    "e-girl": "egirl_eboy",
    "e-boy": "egirl_eboy",
    "cottagecore": "cottagecore",
    "fairycore": "fairycore",
    "chic": "chic",
    "romantic": "romantic",
    # style GT không có trong dataset 49 — map về gần nhất
    "tropical": "casual",
    "relaxed casual": "casual",
    "layered streetwear": "streetwear",
    "minimalist accessories": "minimalist",
    "smart casual": "smart_casual",
    "artsy": "artsy",
}


def _norm_style(s: str) -> str:
    """Normalize tên style GT hoặc AI về style_id chuẩn."""
    key = _norm(s)
    if key in _STYLE_ALIAS:
        return _STYLE_ALIAS[key]
    # fuzzy fallback: tìm alias gần nhất nếu score > 0.8
    best_score = 0.0
    best_id = key
    for alias, sid in _STYLE_ALIAS.items():
        sc = _fuzzy(key, alias)
        if sc > best_score:
            best_score = sc
            best_id = sid
    if best_score >= 0.8:
        return best_id
    return key  # giữ nguyên nếu không khớp


# ---------------------------------------------------------------------------
# §3 — Parse ground truth từ chuỗi nối liền
# ---------------------------------------------------------------------------

def _is_empty_gt(raw: str) -> bool:
    """True nếu chuỗi GT rỗng hoặc chỉ chứa dấu = (placeholder)."""
    s = (raw or "").strip()
    return not s or bool(re.match(r"^=+$", s))


def parse_gt_pairs(raw: str) -> List[Dict[str, str]]:
    """
    Parse chuỗi 'ItemName — StyleName' (nối liền hoặc cách nhau bởi dấu phẩy).
    Hỗ trợ cả dấu '—', '–', '-' (có hoặc không có space xung quanh).

    Trả về list [{'item': ..., 'style': ..., 'style_id': ...}]
    """
    if not raw or not raw.strip():
        return []

    text = raw.replace("—", "—").replace("–", "—").replace(" - ", " — ")

    # Một cặp trên một đoạn (sau tách phẩy): "Flat cap — Classic"
    one_pair = re.compile(
        r"^([A-Z][A-Za-z0-9\s\-\/\(\)\'\.]+?)\s*—\s*([A-Za-z0-9][A-Za-z0-9\s\-\/]*)$",
    )
    # Nhiều cặp nối liền không có phẩy (legacy)
    glued = re.compile(
        r"([A-Z][A-Za-z0-9\s\-\/\(\)\'\.]+?)\s*—\s*([A-Za-z0-9][A-Za-z0-9\s\-\/]+?)(?=(?:[A-Z][a-z0-9])|\Z)"
    )

    results: List[Dict[str, str]] = []
    for seg in re.split(r"\s*,\s*", text):
        seg = seg.strip()
        if not seg:
            continue
        m1 = one_pair.match(seg)
        if m1:
            item = m1.group(1).strip().rstrip("- ")
            style = m1.group(2).strip()
            if len(item) >= 2 and len(style) >= 2:
                results.append({
                    "item": item,
                    "item_norm": _norm(item),
                    "style": style,
                    "style_id": _norm_style(style),
                })
            continue
        for m2 in glued.finditer(seg):
            item = m2.group(1).strip().rstrip("- ")
            style = m2.group(2).strip()
            if len(item) >= 2 and len(style) >= 2:
                results.append({
                    "item": item,
                    "item_norm": _norm(item),
                    "style": style,
                    "style_id": _norm_style(style),
                })
    return results


def parse_gt_overall_style(pairs: List[Dict]) -> str:
    """Phong cách GT tổng thể = style_id xuất hiện nhiều nhất trong danh sách items."""
    if not pairs:
        return ""
    counts = Counter(p["style_id"] for p in pairs)
    return counts.most_common(1)[0][0]


def resolve_gt_overall_for_sample(obj: Dict[str, Any], gt_pairs: List[Dict]) -> Tuple[str, str, str]:
    """
    Trả về (style_id_đã chuẩn hoá, nguồn 'explicit'|'mode', gt_overall_raw nếu explicit).
    Nếu có gt_overall_raw / gt_outfit_overall hợp lệ → dùng explicit; không thì mode trên các món.
    """
    raw = (obj.get("gt_overall_raw") or obj.get("gt_outfit_overall") or "").strip()
    if raw and not _is_empty_gt(raw):
        return _norm_style(raw), "explicit", raw
    return parse_gt_overall_style(gt_pairs), "mode", ""


def overall_style_mode_from_items(items: List[Dict]) -> str:
    """Cùng quy tắc với parse_gt_overall_style: mode của style_id trên các món dự đoán."""
    if not items:
        return ""
    counts: Counter[str] = Counter()
    for it in items:
        sid = (it.get("style_id") or "").strip().lower()
        if sid:
            counts[sid] += 1
    if not counts:
        return ""
    return counts.most_common(1)[0][0]


# Nhóm tương đương (chỉ khi --overall-soft): giảm lệch do ranh giới phong cách gần nhau.
_OVERALL_SOFT_EQUIV_GROUPS: Tuple[frozenset, ...] = (
    frozenset({"casual", "smart_casual", "normcore", "basic", "comfy"}),
    frozenset({"classic", "formal", "business", "old_money", "preppy"}),
    frozenset({"streetwear", "skater_style", "hip_hop", "hypebeast", "techwear", "japanese_streetwear", "korean_style"}),
    frozenset({"elegant", "chic", "romantic", "feminine", "parisian_style"}),
    frozenset({"minimalist", "normcore", "comfy"}),
)


def _overall_style_match(gt: str, pred: str, soft: bool) -> bool:
    g = (gt or "").strip().lower()
    p = (pred or "").strip().lower()
    if not g or not p:
        return False
    if g == p:
        return True
    if not soft:
        return False
    for grp in _OVERALL_SOFT_EQUIV_GROUPS:
        if g in grp and p in grp:
            return True
    return False


def compute_macro_f1_multiclass(y_true: List[str], y_pred: List[str]) -> float:
    """
    Macro F1 theo nhãn style_id: trung bình F1 từng lớp (chỉ các lớp xuất hiện trong y_true hoặc y_pred).
    Mỗi sample = một outfit, nhãn từ Top-1 overall.
    """
    if not y_true or len(y_true) != len(y_pred):
        return 0.0
    labels = sorted(set(y_true) | set(y_pred))
    if not labels:
        return 0.0
    f1s: List[float] = []
    for c in labels:
        tp = sum(1 for t, p in zip(y_true, y_pred) if t == c and p == c)
        fp = sum(1 for t, p in zip(y_true, y_pred) if t != c and p == c)
        fn = sum(1 for t, p in zip(y_true, y_pred) if t == c and p != c)
        pr = tp / (tp + fp) if (tp + fp) else 0.0
        rc = tp / (tp + fn) if (tp + fn) else 0.0
        f1 = (2 * pr * rc / (pr + rc)) if (pr + rc) else 0.0
        f1s.append(f1)
    return round(sum(f1s) / len(f1s), 4) if f1s else 0.0


# ---------------------------------------------------------------------------
# §4 — Gọi API /api/analyze
# ---------------------------------------------------------------------------

def call_analyze_api(
    api_base: str,
    image_path: Path,
    user_id: int,
    timeout_s: int = 120,
) -> Dict[str, Any]:
    """
    Gọi POST /api/analyze với file ảnh và user_id (bắt buộc bởi backend).
    Luôn gửi skip_history=1 để không ghi MySQL history khi chạy benchmark.
    Trả về dict chứa 'status_code', 'json', 'error'.
    """
    url = api_base.rstrip("/") + "/api/analyze"
    try:
        with open(image_path, "rb") as f:
            mime = "image/png" if image_path.suffix.lower() == ".png" else "image/jpeg"
            resp = requests.post(
                url,
                files={"images": (image_path.name, f, mime)},
                data={"user_id": str(int(user_id)), "skip_history": "1"},
                timeout=timeout_s,
            )
        try:
            j = resp.json()
        except Exception:
            j = {}
        api_err = ""
        if resp.status_code not in (200, 201) and isinstance(j, dict):
            api_err = str(j.get("error") or "").strip()
        return {"status_code": resp.status_code, "json": j, "error": api_err}
    except Exception as e:
        return {"status_code": 0, "json": {}, "error": f"{type(e).__name__}: {e}"}


def extract_pred_from_response(resp_json: Dict[str, Any]) -> Dict[str, Any]:
    """
    Trích thông tin từ response API:
    - items: list {'item', 'item_norm', 'style_id', 'confidence'}
    - overall_style_id
    - overall_style_top_k: list style_id đã _norm_style (Top-k outfit cùng nguồn aggregate_styles)
    """
    results = resp_json.get("results") or []
    if not results:
        return {"items": [], "overall_style_id": "", "overall_style_top_k": []}

    r0 = results[0] if isinstance(results[0], dict) else {}
    raw_items = r0.get("items") or []
    overall_style = _norm_style(str(r0.get("overall_style") or ""))

    top_k_disp = r0.get("overall_style_top_k")
    overall_top_k_ids: List[str] = []
    if isinstance(top_k_disp, list) and top_k_disp:
        for x in top_k_disp:
            overall_top_k_ids.append(_norm_style(str(x)))
    if not overall_top_k_ids and overall_style:
        overall_top_k_ids = [overall_style]

    items = []
    for it in raw_items:
        if not isinstance(it, dict):
            continue
        name = it.get("item") or it.get("item_en") or it.get("item_type") or ""
        style_raw = it.get("final_style") or it.get("style") or ""
        conf = it.get("confidence")
        try:
            conf_f = float(conf) if conf is not None else 0.7
            if conf_f > 1.0:
                conf_f /= 100.0
        except (TypeError, ValueError):
            conf_f = 0.7
        items.append({
            "item": str(name).strip(),
            "item_norm": _norm(str(name)),
            "style_id": _norm_style(str(style_raw)),
            "confidence": round(conf_f, 3),
        })

    return {"items": items, "overall_style_id": overall_style, "overall_style_top_k": overall_top_k_ids}


# ---------------------------------------------------------------------------
# §5 — Matching item: fuzzy + exact
# ---------------------------------------------------------------------------

def match_items(
    gt_items: List[Dict],
    pred_items: List[Dict],
    threshold: float = 0.6,
) -> Tuple[List[Tuple], List[Dict], List[Dict]]:
    """
    Greedy fuzzy matching: ghép item GT với item pred có similarity cao nhất.
    threshold: ngưỡng tối thiểu để coi là "khớp".

    Trả về:
    - matched: list (gt_item, pred_item, similarity)
    - unmatched_gt: gt items không khớp được (false negatives)
    - unmatched_pred: pred items không khớp được (false positives)
    """
    used_pred = set()
    matched = []
    unmatched_gt = []

    for gi, g in enumerate(gt_items):
        best_score = 0.0
        best_pi = -1
        for pi, p in enumerate(pred_items):
            if pi in used_pred:
                continue
            sc = _fuzzy(g["item_norm"], p["item_norm"])
            if sc > best_score:
                best_score = sc
                best_pi = pi
        if best_pi >= 0 and best_score >= threshold:
            matched.append((g, pred_items[best_pi], round(best_score, 3)))
            used_pred.add(best_pi)
        else:
            unmatched_gt.append(g)

    unmatched_pred = [p for pi, p in enumerate(pred_items) if pi not in used_pred]
    return matched, unmatched_gt, unmatched_pred


def compute_item_f1(
    matched: List[Tuple],
    unmatched_gt: List[Dict],
    unmatched_pred: List[Dict],
) -> Tuple[float, float, float]:
    tp = len(matched)
    fp = len(unmatched_pred)
    fn = len(unmatched_gt)
    precision = tp / (tp + fp) if (tp + fp) else 0.0
    recall = tp / (tp + fn) if (tp + fn) else 0.0
    f1 = (2 * precision * recall / (precision + recall)) if (precision + recall) else 0.0
    return round(precision, 4), round(recall, 4), round(f1, 4)


def count_style_correct(matched: List[Tuple]) -> int:
    """Số cặp (gt, pred) đã khớp item mà style_id cũng khớp."""
    return sum(1 for g, p, _ in matched if g["style_id"] == p["style_id"])


def compute_style_accuracy(matched: List[Tuple]) -> float:
    """Trong các cặp (gt, pred) đã khớp item, tỷ lệ style_id giống nhau."""
    if not matched:
        return 0.0
    return round(count_style_correct(matched) / len(matched), 4)


# ---------------------------------------------------------------------------
# §6 — Main evaluation loop
# ---------------------------------------------------------------------------

# Bố cục CSV: một sample = một dòng vật lý (không \n trong ô); metric trước, chi tiết dài cuối — mở Excel dễ quét.
REPORT_CSV_FIELDNAMES = [
    "sample_id",
    "image_name",
    "f1",
    "precision",
    "recall",
    "overall_style_hit",
    "style_item_accuracy",
    "tp",
    "style_correct_count",
    "fp",
    "fn",
    "gt_item_count",
    "pred_item_count",
    "gt_overall_style",
    "gt_overall_raw",
    "gt_overall_source",
    "pred_overall_from_api",
    "pred_overall_mode",
    "status_code",
    "api_error",
    "items_gt",
    "items_pred",
    "matched_pairs_detail",
    "unmatched_gt",
    "unmatched_pred",
    "image_file",
]


def _fmt_items_one_line(items: List[str], sep: str = " | ") -> str:
    if not items:
        return ""
    out: List[str] = []
    for x in items:
        s = str(x).strip()
        if s:
            out.append(s.replace("\n", " ").replace(sep.strip(), " "))
    return sep.join(out)


def _fmt_matched_one_line(pairs: List[Dict[str, Any]]) -> str:
    if not pairs:
        return ""
    parts: List[str] = []
    for d in pairs:
        sty = "Y" if d.get("style_match") else "N"
        parts.append(
            f"{d.get('gt', '')}→{d.get('pred', '')} sim={d.get('sim', '')} "
            f"gt_st={d.get('gt_style', '')} pr_st={d.get('pred_style', '')} ok={sty}"
        )
    return " ; ".join(parts)


def report_row_to_csv_flat(row: Dict[str, Any]) -> Dict[str, str]:
    """Chuyển một hàng báo cáo sang ô CSV một dòng (danh sách ghép bằng ' | ')."""
    gt = row.get("gt_items")
    pr = row.get("pred_items")
    if not isinstance(gt, list):
        gt = []
    if not isinstance(pr, list):
        pr = []
    mp = row.get("matched_pairs")
    if not isinstance(mp, list):
        mp = []
    ug = row.get("unmatched_gt")
    up = row.get("unmatched_pred")
    if not isinstance(ug, list):
        ug = []
    if not isinstance(up, list):
        up = []

    def _scalar(k: str) -> str:
        v = row.get(k)
        if v is None:
            return ""
        return str(v)

    raw_img = _scalar("image_file")
    image_name = ""
    if raw_img:
        try:
            image_name = Path(str(raw_img).strip()).name
        except OSError:
            image_name = ""

    return {
        "sample_id": _scalar("sample_id"),
        "image_name": image_name,
        "f1": _scalar("f1"),
        "precision": _scalar("precision"),
        "recall": _scalar("recall"),
        "overall_style_hit": _scalar("overall_style_hit"),
        "style_item_accuracy": _scalar("style_item_accuracy"),
        "tp": _scalar("tp"),
        "style_correct_count": _scalar("style_correct_count"),
        "fp": _scalar("fp"),
        "fn": _scalar("fn"),
        "gt_item_count": _scalar("gt_item_count"),
        "pred_item_count": _scalar("pred_item_count"),
        "gt_overall_style": _scalar("gt_overall_style"),
        "gt_overall_raw": _scalar("gt_overall_raw"),
        "gt_overall_source": _scalar("gt_overall_source"),
        "pred_overall_from_api": _scalar("pred_overall_from_api"),
        "pred_overall_mode": _scalar("pred_overall_mode"),
        "status_code": _scalar("status_code"),
        "api_error": _scalar("api_error"),
        "items_gt": _fmt_items_one_line([str(x) for x in gt]),
        "items_pred": _fmt_items_one_line([str(x) for x in pr]),
        "matched_pairs_detail": _fmt_matched_one_line(mp),
        "unmatched_gt": _fmt_items_one_line([str(x) for x in ug]),
        "unmatched_pred": _fmt_items_one_line([str(x) for x in up]),
        "image_file": raw_img,
    }


_SUMMARY_LABELS_VI = {
    "run_id": "Mã lần chạy",
    "saved_at": "Thời gian lưu",
    "samples_total": "Tổng sample trong GT",
    "samples_api_attempted": "Số sample đã gọi API",
    "samples_scored": "Số sample tính metric",
    "fuzzy_threshold": "Ngưỡng fuzzy match",
    "avg_precision": "Precision trung bình (item)",
    "avg_recall": "Recall trung bình (item)",
    "avg_f1": "F1 trung bình (item)",
    "items_correct_total": "Tổng item nhận diện đúng (TP)",
    "items_style_correct_total": "Tổng item có phong cách đúng",
    "gt_items_total": "Tổng món ground truth",
    "pred_items_total": "Tổng món AI dự đoán",
    "style_per_item_accuracy": "Tỷ lệ style đúng / TP",
    "matched_pairs_total": "Tổng cặp item khớp (alias TP)",
    "overall_style_accuracy": "Độ chính xác overall style",
    "overall_top1_accuracy": "Outfit Top-1 accuracy",
    "overall_top3_accuracy": "Outfit Top-3 accuracy",
    "overall_macro_f1": "Outfit Macro F1",
    "overall_style_samples_scored": "Số sample chấm overall",
    "overall_match": "Chế độ so overall",
    "overall_soft": "Overall soft match",
    "output_dir": "Thư mục kết quả",
}


def save_report_xlsx(
    report_rows: List[Dict[str, Any]],
    summary: Dict[str, Any],
    xlsx_path: Path,
) -> None:
    """Ghi báo cáo eval ra file Excel (.xlsx): sheet Chi tiết + Tổng hợp."""
    if Workbook is None:
        raise RuntimeError("Thiếu openpyxl. Cài: pip install openpyxl")

    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = "Chi tiết"

    flat_rows = [report_row_to_csv_flat(row) for row in report_rows]
    headers = list(REPORT_CSV_FIELDNAMES)
    ws_detail.append(headers)
    for cell in ws_detail[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for flat in flat_rows:
        ws_detail.append([flat.get(h, "") for h in headers])

    ws_detail.freeze_panes = "A2"
    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        max_len = len(header)
        for row in flat_rows:
            max_len = max(max_len, len(str(row.get(header, ""))))
        ws_detail.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)

    wrap_cols = {
        "items_gt", "items_pred", "matched_pairs_detail",
        "unmatched_gt", "unmatched_pred", "image_file",
    }
    wrap_indices = {headers.index(c) + 1 for c in wrap_cols if c in headers}
    for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row):
        for idx in wrap_indices:
            row[idx - 1].alignment = Alignment(wrap_text=True, vertical="top")

    ws_sum = wb.create_sheet("Tổng hợp")
    ws_sum.append(["Chỉ số", "Giá trị"])
    ws_sum["A1"].font = Font(bold=True)
    ws_sum["B1"].font = Font(bold=True)
    for key, value in summary.items():
        label = _SUMMARY_LABELS_VI.get(key, key)
        ws_sum.append([label, value])
    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 28

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def _resolve_eval_image_path(obj: Dict[str, Any], media_dir: Path) -> Optional[Path]:
    """
    Tìm file ảnh cho một sample:
    1) image_file trong JSON nếu file còn tồn tại;
    2) media_dir/<image_id>.png|jpg (benchmark Excel: ST001, ST002…);
    3) media_dir/<suffix sau _ trong sample_id> (vd. 0001_ST001 → ST001);
    4) media_dir/<imageN> theo pattern cũ (image1, image2…).
    """
    _EXTS = (".png", ".jpg", ".jpeg", ".webp")

    def _first_existing(stems: List[str]) -> Optional[Path]:
        for stem in stems:
            s = (stem or "").strip()
            if not s:
                continue
            for ext in _EXTS:
                candidate = (media_dir / (s + ext)).resolve()
                if candidate.is_file():
                    return candidate
        return None

    raw = obj.get("image_file") or ""
    if raw:
        p = Path(str(raw).strip()).expanduser()
        try:
            p = p.resolve()
        except OSError:
            pass
        if p.is_file():
            return p
        # Đường dẫn tuyệt đối cũ (copy project) — thử lại theo tên file trong media_dir
        by_basename = _first_existing([p.stem])
        if by_basename:
            return by_basename

    image_id = str(obj.get("image_id") or "").strip()
    if image_id:
        found = _first_existing([image_id])
        if found:
            return found

    sid = str(obj.get("sample_id") or "").strip()
    if "_" in sid:
        found = _first_existing([sid.split("_", 1)[1]])
        if found:
            return found

    img_name_match = re.search(r"(image\d+)", sid, re.IGNORECASE)
    if img_name_match:
        found = _first_existing([img_name_match.group(1)])
        if found:
            return found

    return None


def main():
    ap = argparse.ArgumentParser(
        description="Kiểm chứng độ chính xác hệ thống nhận dạng thời trang."
    )
    ap.add_argument(
        "--gt",
        default="artifacts/eval_excel/ground_truth_parsed.jsonl",
        help="Đường dẫn tới ground_truth_parsed.jsonl",
    )
    ap.add_argument(
        "--media",
        default="artifacts/eval_excel/media",
        help="Thư mục chứa file ảnh benchmark (vd: ST001.png)",
    )
    ap.add_argument(
        "--api",
        default="http://127.0.0.1:5000",
        help="Base URL của Flask server (vd: http://127.0.0.1:5000)",
    )
    ap.add_argument(
        "--user-id",
        type=int,
        default=None,
        help="ID user trong DB (bắt buộc khi gọi API). Nên dùng user admin để không trừ lượt.",
    )
    ap.add_argument(
        "--out",
        default="artifacts/eval_excel/results",
        help="Thư mục lưu kết quả",
    )
    ap.add_argument(
        "--limit",
        type=int,
        default=0,
        help="Chỉ chạy N ảnh đầu (0 = tất cả)",
    )
    ap.add_argument(
        "--sleep",
        type=float,
        default=1.0,
        help="Số giây nghỉ giữa các lần gọi API",
    )
    ap.add_argument(
        "--timeout",
        type=int,
        default=120,
        help="Timeout mỗi lần gọi API (giây)",
    )
    ap.add_argument(
        "--fuzzy-threshold",
        type=float,
        default=0.60,
        help="Ngưỡng fuzzy match item (0..1, mặc định 0.60)",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Chỉ parse GT, không gọi API (kiểm tra GT parser)",
    )
    ap.add_argument(
        "--overall-match",
        choices=("mode", "api"),
        default="mode",
        help="So overall với GT: mode = mode(style_id các món pred), khớp định nghĩa GT (mặc định). api = dùng overall_style từ backend.",
    )
    ap.add_argument(
        "--overall-soft",
        action="store_true",
        help="Coi một số cặp phong cách gần nghĩa là khớp (chỉ ảnh hưởng overall_style_hit).",
    )
    args = ap.parse_args()

    # ── Setup ──────────────────────────────────────────────────────────────
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

    gt_path = Path(args.gt).expanduser().resolve()
    media_dir = Path(args.media).expanduser().resolve()
    out_dir = Path(args.out).expanduser().resolve()
    out_dir.mkdir(parents=True, exist_ok=True)

    if not gt_path.exists():
        print(f"[ERROR] Không tìm thấy GT file: {gt_path}", file=sys.stderr)
        sys.exit(1)

    # ── Load GT ───────────────────────────────────────────────────────────
    samples_raw = []
    with open(gt_path, encoding="utf-8-sig") as f:
        for line in f:
            line = line.strip()
            if line:
                try:
                    samples_raw.append(json.loads(line))
                except json.JSONDecodeError:
                    pass

    if args.limit > 0:
        samples_raw = samples_raw[:args.limit]

    print(f"[INFO] Loaded {len(samples_raw)} samples từ GT", file=sys.stderr)

    if not args.dry_run and args.user_id is None:
        print(
            "[ERROR] Thiếu --user-id. API /api/analyze yêu cầu đăng nhập (user_id). "
            "Ví dụ: --user-id 1 (user admin trong DB).",
            file=sys.stderr,
        )
        sys.exit(2)

    # ── Dry run: chỉ kiểm tra GT parser ───────────────────────────────────
    if args.dry_run:
        print("\n=== DRY RUN: kiểm tra GT parser ===")
        total_pairs = 0
        for obj in samples_raw:
            raw_i = obj.get("gt_items_raw", "")
            raw_s = obj.get("gt_style_raw", "")
            effective = raw_i
            if _is_empty_gt(raw_i) and not _is_empty_gt(raw_s):
                effective = raw_s
            pairs = parse_gt_pairs(effective)
            total_pairs += len(pairs)
            gt_overall_style, gt_overall_source, gt_overall_raw = resolve_gt_overall_for_sample(obj, pairs)
            print(f"\n[{obj['sample_id']}] — {len(pairs)} items — overall_GT=({gt_overall_source}) {gt_overall_style}" + (f'  [raw: {gt_overall_raw}]' if gt_overall_raw else ''))
            for p in pairs:
                print(f"  '{p['item']}' → style_id={p['style_id']}")
        print(f"\nTổng cộng: {total_pairs} item-style pairs")
        return

    # ── Eval loop ─────────────────────────────────────────────────────────
    report_rows = []
    error_log = []

    # Tổng hợp metric
    sum_p = sum_r = sum_f1 = 0.0
    sum_style_item = 0.0          # style accuracy per item (matched)
    sum_style_overall = 0         # overall style hit
    n_style_overall_scored = 0
    n_matched_pairs_total = 0
    n_items_style_correct_total = 0
    n_gt_items_total = 0
    n_pred_items_total = 0
    n_samples_scored = 0
    n_samples_api_attempted = 0
    sum_overall_top1 = 0
    sum_overall_top3 = 0
    macro_y_true: List[str] = []
    macro_y_pred: List[str] = []

    for idx, obj in enumerate(samples_raw, start=1):
        sample_id = obj["sample_id"]
        gt_items_raw = obj.get("gt_items_raw", "")
        gt_style_raw = obj.get("gt_style_raw", "")

        # Parse GT — nếu gt_items_raw rỗng/placeholder thì fallback sang gt_style_raw
        effective_raw = gt_items_raw
        if _is_empty_gt(gt_items_raw) and not _is_empty_gt(gt_style_raw):
            effective_raw = gt_style_raw
        gt_pairs = parse_gt_pairs(effective_raw)
        gt_overall_style, gt_overall_source, gt_overall_raw = resolve_gt_overall_for_sample(obj, gt_pairs)

        img_file = _resolve_eval_image_path(obj, media_dir)

        if img_file is None:
            error_log.append(f"[{sample_id}] Không tìm thấy file ảnh trong {media_dir}")
            print(f"[{idx}/{len(samples_raw)}] {sample_id}: SKIP (no image)", file=sys.stderr)
            continue

        # Gọi API
        n_samples_api_attempted += 1
        print(f"[{idx}/{len(samples_raw)}] {sample_id}: gọi API...", end=" ", file=sys.stderr)
        resp = call_analyze_api(args.api, img_file, args.user_id, args.timeout)
        status_code = resp["status_code"]
        api_error = resp["error"]

        if api_error or status_code not in (200, 201):
            err_msg = api_error or f"HTTP {status_code}"
            error_log.append(f"[{sample_id}] API error: {err_msg}")
            print(f"ERROR: {err_msg}", file=sys.stderr)
            report_rows.append({
                "sample_id": sample_id,
                "image_file": str(img_file),
                "status_code": status_code,
                "api_error": err_msg,
                "gt_item_count": len(gt_pairs),
                "pred_item_count": 0,
                "precision": 0.0, "recall": 0.0, "f1": 0.0,
                "tp": 0, "fp": 0, "fn": len(gt_pairs),
                "style_correct_count": 0,
                "style_item_accuracy": 0.0,
                "gt_overall_raw": gt_overall_raw,
                "gt_overall_source": gt_overall_source,
                "gt_overall_style": gt_overall_style,
                "pred_overall_mode": "",
                "pred_overall_from_api": "",
                "overall_style_hit": 0,
                "gt_items": [p["item"] for p in gt_pairs],
                "pred_items": [],
                "matched_pairs": [],
                "unmatched_gt": [p["item"] for p in gt_pairs],
                "unmatched_pred": [],
            })
            n_gt_items_total += len(gt_pairs)
            if args.sleep > 0:
                time.sleep(args.sleep)
            continue

        pred = extract_pred_from_response(resp["json"])
        pred_items = pred["items"]
        pred_overall_api = pred["overall_style_id"]
        pred_overall_mode = overall_style_mode_from_items(pred_items)
        pred_for_overall = pred_overall_api if args.overall_match == "api" else pred_overall_mode

        # Match items
        matched, unmatched_gt, unmatched_pred = match_items(
            gt_pairs, pred_items, threshold=args.fuzzy_threshold
        )
        precision, recall, f1 = compute_item_f1(matched, unmatched_gt, unmatched_pred)
        style_item_acc = compute_style_accuracy(matched)
        style_correct_count = count_style_correct(matched)

        # Overall style hit (GT = mode item GT; pred = mode item hoặc API tùy --overall-match)
        style_hit = 0
        if gt_overall_style:
            style_hit = 1 if _overall_style_match(gt_overall_style, pred_for_overall, args.overall_soft) else 0
            n_style_overall_scored += 1
            sum_style_overall += style_hit

            # Outfit Top-1 / Top-3 / Macro F1: luôn dùng xếp hạng aggregate_styles (field overall_style_top_k)
            tk = pred.get("overall_style_top_k") or []
            if not tk and pred.get("overall_style_id"):
                tk = [pred["overall_style_id"]]
            if not tk:
                tk = [""]
            sum_overall_top1 += 1 if _overall_style_match(gt_overall_style, tk[0], args.overall_soft) else 0
            sum_overall_top3 += (
                1
                if any(
                    _overall_style_match(gt_overall_style, t, args.overall_soft)
                    for t in tk[:3]
                )
                else 0
            )
            macro_y_true.append(gt_overall_style)
            macro_y_pred.append(tk[0] if tk else "")

        # Accumulate
        sum_p += precision
        sum_r += recall
        sum_f1 += f1
        sum_style_item += style_item_acc * len(matched)
        n_matched_pairs_total += len(matched)
        n_items_style_correct_total += style_correct_count
        n_gt_items_total += len(gt_pairs)
        n_pred_items_total += len(pred_items)
        n_samples_scored += 1

        tp = len(matched)
        fp = len(unmatched_pred)
        fn = len(unmatched_gt)

        print(
            f"P={precision:.2f} R={recall:.2f} F1={f1:.2f} "
            f"tp={tp} style_ok={style_correct_count} fp={fp} fn={fn} "
            f"style_item={style_item_acc:.2f} overall={'HIT' if style_hit else 'MISS'} "
            f"(gt={gt_overall_style} pred={pred_for_overall})",
            file=sys.stderr,
        )

        matched_pairs_data = [
            {
                "gt": g["item"],
                "pred": p["item"],
                "sim": s,
                "gt_style": g["style_id"],
                "pred_style": p["style_id"],
                "style_match": bool(g["style_id"] == p["style_id"]),
            }
            for g, p, s in matched
        ]
        report_rows.append({
            "sample_id": sample_id,
            "image_file": str(img_file),
            "status_code": status_code,
            "api_error": "",
            "gt_item_count": len(gt_pairs),
            "pred_item_count": len(pred_items),
            "precision": precision,
            "recall": recall,
            "f1": f1,
            "tp": tp,
            "fp": fp,
            "fn": fn,
            "style_correct_count": style_correct_count,
            "style_item_accuracy": style_item_acc,
            "gt_overall_raw": gt_overall_raw,
            "gt_overall_source": gt_overall_source,
            "gt_overall_style": gt_overall_style,
            "pred_overall_mode": pred_overall_mode,
            "pred_overall_from_api": pred_overall_api,
            "overall_style_hit": style_hit,
            "gt_items": [p["item"] for p in gt_pairs],
            "pred_items": [p["item"] for p in pred_items],
            "matched_pairs": matched_pairs_data,
            "unmatched_gt": [p["item"] for p in unmatched_gt],
            "unmatched_pred": [p["item"] for p in unmatched_pred],
        })

        if args.sleep > 0:
            time.sleep(args.sleep)

    # ── Summary ───────────────────────────────────────────────────────────
    n = n_samples_scored
    summary = {
        "samples_total": len(samples_raw),
        "samples_api_attempted": n_samples_api_attempted,
        "samples_scored": n,
        "fuzzy_threshold": args.fuzzy_threshold,

        # Item detection
        "avg_precision": round(sum_p / n, 4) if n else 0.0,
        "avg_recall": round(sum_r / n, 4) if n else 0.0,
        "avg_f1": round(sum_f1 / n, 4) if n else 0.0,
        "items_correct_total": n_matched_pairs_total,
        "items_style_correct_total": n_items_style_correct_total,
        "gt_items_total": n_gt_items_total,
        "pred_items_total": n_pred_items_total,

        # Per-item style (trong các cặp đã match)
        "style_per_item_accuracy": round(sum_style_item / n_matched_pairs_total, 4)
            if n_matched_pairs_total else 0.0,
        "matched_pairs_total": n_matched_pairs_total,

        # Overall style
        "overall_style_accuracy": round(sum_style_overall / n_style_overall_scored, 4)
            if n_style_overall_scored else 0.0,
        "overall_top1_accuracy": round(sum_overall_top1 / n_style_overall_scored, 4)
            if n_style_overall_scored else 0.0,
        "overall_top3_accuracy": round(sum_overall_top3 / n_style_overall_scored, 4)
            if n_style_overall_scored else 0.0,
        "overall_macro_f1": compute_macro_f1_multiclass(macro_y_true, macro_y_pred)
            if macro_y_true else 0.0,
        "overall_style_samples_scored": n_style_overall_scored,
        "overall_match": args.overall_match,
        "overall_soft": bool(args.overall_soft),

        "output_dir": str(out_dir),
    }

    # ── Write outputs ─────────────────────────────────────────────────────
    now = datetime.now()
    stamp = now.strftime("%Y%m%d_%H%M%S") + f"_{now.microsecond:06d}"
    summary["run_id"] = stamp
    summary["saved_at"] = now.isoformat(timespec="seconds")

    if report_rows:
        for csv_path in (out_dir / f"report_{stamp}.csv", out_dir / "report.csv"):
            with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.DictWriter(
                    f,
                    fieldnames=REPORT_CSV_FIELDNAMES,
                    quoting=csv.QUOTE_MINIMAL,
                )
                writer.writeheader()
                for raw_row in report_rows:
                    writer.writerow(report_row_to_csv_flat(raw_row))
        print(
            f"[INFO] CSV saved: {out_dir / 'report.csv'} (latest), "
            f"{out_dir / f'report_{stamp}.csv'} (archive)",
            file=sys.stderr,
        )

        for xlsx_path in (out_dir / f"report_{stamp}.xlsx", out_dir / "report.xlsx"):
            try:
                save_report_xlsx(report_rows, summary, xlsx_path)
            except RuntimeError as exc:
                print(f"[WARN] Không ghi Excel: {exc}", file=sys.stderr)
                break
        else:
            print(
                f"[INFO] Excel saved: {out_dir / 'report.xlsx'} (latest), "
                f"{out_dir / f'report_{stamp}.xlsx'} (archive)",
                file=sys.stderr,
            )

        for jsonl_path in (out_dir / f"report_{stamp}.jsonl", out_dir / "report.jsonl"):
            with open(jsonl_path, "w", encoding="utf-8") as f:
                for row in report_rows:
                    f.write(json.dumps(row, ensure_ascii=False) + "\n")
        print(
            f"[INFO] JSONL saved: {out_dir / 'report.jsonl'} (latest), "
            f"{out_dir / f'report_{stamp}.jsonl'} (archive)",
            file=sys.stderr,
        )

    summary_text = json.dumps(summary, ensure_ascii=False, indent=2)

    summary_hist_path = out_dir / f"summary_{stamp}.json"
    summary_hist_path.write_text(summary_text, encoding="utf-8")
    print(f"[INFO] Summary (archive): {summary_hist_path}", file=sys.stderr)

    summary_path = out_dir / "summary.json"
    summary_path.write_text(summary_text, encoding="utf-8")
    print(f"[INFO] Summary (latest) : {summary_path}", file=sys.stderr)

    err_body = "\n".join(error_log)
    err_hist = out_dir / f"errors_{stamp}.txt"
    err_latest = out_dir / "errors.txt"
    err_hist.write_text(err_body, encoding="utf-8")
    err_latest.write_text(err_body, encoding="utf-8")
    if error_log:
        print(
            f"[WARN] {len(error_log)} errors: {err_latest} (archive: {err_hist})",
            file=sys.stderr,
        )
    else:
        print(
            f"[INFO] No errors logged: {err_latest} (archive: {err_hist})",
            file=sys.stderr,
        )

    # ── Print summary ─────────────────────────────────────────────────────
    print("\n" + "=" * 55)
    print("  KẾT QUẢ KIỂM CHỨNG")
    print("=" * 55)
    print(f"  Samples gọi API     : {summary['samples_api_attempted']}/{len(samples_raw)}")
    print(f"  Samples tính metric : {n}/{len(samples_raw)}  (API thành công)")
    print(f"  Fuzzy threshold     : {args.fuzzy_threshold}")
    print(f"  Overall match       : {args.overall_match}  (soft={args.overall_soft})")
    print()
    print("  [Item Detection]")
    print(f"  Avg Precision       : {summary['avg_precision']:.4f}")
    print(f"  Avg Recall          : {summary['avg_recall']:.4f}")
    print(f"  Avg F1              : {summary['avg_f1']:.4f}")
    print(f"  Item đúng (TP)      : {summary['items_correct_total']} / {summary['gt_items_total']} GT")
    if summary['gt_items_total']:
        item_recall_pct = 100.0 * summary['items_correct_total'] / summary['gt_items_total']
        print(f"                        ({item_recall_pct:.1f}% recall tổng)")
    print(f"  Pred items total    : {summary['pred_items_total']}")
    print()
    print("  [Style Accuracy]")
    print(f"  Style đúng (item)   : {summary['items_style_correct_total']} / "
          f"{summary['items_correct_total']} TP")
    print(f"  Per-item style acc  : {summary['style_per_item_accuracy']:.4f}  "
          f"(= style đúng / TP)")
    print(f"  Overall style acc   : {summary['overall_style_accuracy']:.4f}  "
          f"(trên {summary['overall_style_samples_scored']} sample)  "
          f"[theo --overall-match]")
    print(f"  Outfit Top-1 acc    : {summary['overall_top1_accuracy']:.4f}  "
          f"(xếp hạng aggregate_styles, field overall_style_top_k)")
    print(f"  Outfit Top-3 acc    : {summary['overall_top3_accuracy']:.4f}")
    print(f"  Outfit Macro F1     : {summary['overall_macro_f1']:.4f}  (Top-1 mỗi sample)")
    print("=" * 55)


# ---------------------------------------------------------------------------
# §7 — Entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    main()